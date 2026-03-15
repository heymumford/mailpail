# Copyright (C) 2026+ Eric C. Mumford <eric@mumfordengineering.com>
# SPDX-License-Identifier: GPL-3.0-or-later

from __future__ import annotations

import logging
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from aol_email_exporter.models import EmailRecord, ExportConfig, ExportResult

logger = logging.getLogger(__name__)

_HEADERS = ["Date", "From", "To", "CC", "Subject", "Body", "Folder", "Attachments", "Message-ID"]
_MAX_COL_WIDTH = 50
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?\[\]:]")


def _record_to_row(rec: EmailRecord) -> list[object]:
    return [
        rec.date.isoformat() if rec.date else "",
        rec.sender,
        rec.to,
        rec.cc,
        rec.subject,
        rec.body_text,
        rec.folder,
        "Yes" if rec.has_attachments else "No",
        rec.message_id,
    ]


def _auto_width(ws: Worksheet) -> None:
    """Set column widths based on content, capped at _MAX_COL_WIDTH."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, _MAX_COL_WIDTH)


def _sanitize_sheet_name(name: str) -> str:
    """Ensure the sheet name conforms to Excel limits (31 chars, no special chars)."""
    clean = _INVALID_SHEET_CHARS.sub("_", name)
    return clean[:31] if clean else "Sheet"


def _write_sheet(ws: Worksheet, records: list[EmailRecord]) -> None:
    """Write header + data rows to a worksheet, then style it."""
    ws.append(_HEADERS)
    for rec in records:
        ws.append(_record_to_row(rec))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_width(ws)


class ExcelExporter:
    """Export all emails to a single-worksheet Excel file."""

    def export(self, records: list[EmailRecord], config: ExportConfig) -> ExportResult:
        out_dir = Path(config.output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{config.filename_prefix}.xlsx"

        try:
            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.title = "Emails"

            _write_sheet(ws, records)
            wb.save(str(out_path))

            logger.info("Excel export complete: %d records -> %s", len(records), out_path)
            return ExportResult(
                format_name="excel",
                file_path=str(out_path),
                record_count=len(records),
                success=True,
            )
        except Exception as exc:  # noqa: BLE001
            logger.error("Excel export failed: %s", exc)
            return ExportResult(
                format_name="excel",
                file_path=str(out_path),
                record_count=0,
                success=False,
                error=str(exc),
            )


class ExcelSheetsExporter:
    """Export emails to an Excel file with one worksheet per group."""

    def export(self, records: list[EmailRecord], config: ExportConfig) -> ExportResult:
        out_dir = Path(config.output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        group_by = config.excel_group_by or "folder"
        out_path = out_dir / f"{config.filename_prefix}_by_{group_by}.xlsx"

        try:
            groups = self._group_records(records, group_by)
            wb = Workbook()
            # Remove the default empty sheet.
            default_ws = wb.active
            assert default_ws is not None

            first = True
            for name, group_records in sorted(groups.items()):
                sheet_name = _sanitize_sheet_name(name)
                if first:
                    default_ws.title = sheet_name
                    ws = default_ws
                    first = False
                else:
                    ws = wb.create_sheet(title=sheet_name)
                _write_sheet(ws, group_records)

            wb.save(str(out_path))
            logger.info(
                "Excel (grouped by %s) export complete: %d records, %d sheets -> %s",
                group_by,
                len(records),
                len(groups),
                out_path,
            )
            return ExportResult(
                format_name="excel-sheets",
                file_path=str(out_path),
                record_count=len(records),
                success=True,
            )
        except Exception as exc:  # noqa: BLE001
            logger.error("Excel sheets export failed: %s", exc)
            return ExportResult(
                format_name="excel-sheets",
                file_path=str(out_path),
                record_count=0,
                success=False,
                error=str(exc),
            )

    @staticmethod
    def _group_records(records: list[EmailRecord], group_by: str) -> dict[str, list[EmailRecord]]:
        groups: dict[str, list[EmailRecord]] = defaultdict(list)

        for rec in records:
            if group_by == "folder":
                key = rec.folder or "Unknown"
            elif group_by == "sender":
                key = rec.sender or "Unknown"
            elif group_by == "date":
                key = rec.date.strftime("%Y-%m") if rec.date else "Unknown"
            else:
                raise ValueError(f"Unsupported group_by value: {group_by}")
            groups[key].append(rec)

        return dict(groups)
